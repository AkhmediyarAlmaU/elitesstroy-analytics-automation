import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def generate_styled_excel(df: pd.DataFrame, output_path: str) -> None:
    """
    Экспортирует данные в Excel и применяет корпоративное форматирование.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Анализ затрат"
    
    # Включение сетки
    ws.views.sheetView[0].showGridLines = True
    
    # Перенос данных из Pandas в Excel-лист
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
        
    # Стилизация шапки таблицы
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
    # Форматирование ячеек с данными
    data_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center')
    left_align = Alignment(horizontal='left')
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = data_font
            if cell.column in [1, 4]:  # ID и количество счетов по центру
                cell.alignment = center_align
            elif cell.column == 2:     # Название объекта по левому краю
                cell.alignment = left_align
            else:                      # Финансовые показатели
                cell.number_format = '#,##0.00'
                
    # Автоматическое определение ширины столбцов
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(output_path)
    print(f"Отчет успешно сохранен и оформлен: {output_path}")
